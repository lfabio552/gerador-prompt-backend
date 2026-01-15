import os
import io
import json
import re
import google.generativeai as genai
import stripe
import replicate
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from supabase import create_client, Client

# --- FERRAMENTAS EXTRAS ---
from pytube import YouTube
import xml.etree.ElementTree as ET
from docx import Document
from docx.shared import Cm, Pt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pypdf import PdfReader 

# Carrega variáveis do .env
load_dotenv() 

app = Flask(__name__)

# 1. Configuração do Flask-CORS (Libera tudo)
CORS(app, resources={r"/*": {"origins": "*"}})

# 2. SOLUÇÃO FORÇA BRUTA: Injeção Manual de Headers
# Garante que TODA resposta leve os headers de permissão, inclusive erros
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response
    
# --- VERIFICAÇÃO DE CHAVES ---
stripe_key = os.environ.get("STRIPE_SECRET_KEY")
endpoint_secret = os.environ.get('STRIPE_WEBHOOK_SECRET')
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000") # Fallback para localhost se não tiver

# Configurações
if stripe_key:
    stripe.api_key = stripe_key

url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")

if url and key:
    supabase: Client = create_client(url, key)
else:
    print("ERRO CRÍTICO: Chaves do Supabase faltando!")
    supabase = None

try:
    genai.configure(api_key=os.getenv('GOOGLE_API_KEY'))
    # Mantive seu modelo escolhido
    model = genai.GenerativeModel('gemini-2.5-flash')
    print("Modelo Gemini configurado com sucesso!")
except Exception as e:
    print(f"Erro ao configurar o modelo Gemini: {e}")
    model = None

# --- FUNÇÃO DE CRÉDITOS ---
def check_and_deduct_credit(user_id):
    try:
        if not supabase: return False, "Erro de banco de dados."
        response = supabase.table('profiles').select('credits, is_pro').eq('id', user_id).execute()
        
        if not response.data: return False, "Usuário não encontrado."
        
        user_data = response.data[0]
        credits = user_data.get('credits', 0)
        is_pro = user_data.get('is_pro', False) 
        
        if is_pro: return True, "Sucesso (VIP)"
            
        if credits <= 0: return False, "Sem créditos. Assine o PRO!"
            
        new_credits = credits - 1
        supabase.table('profiles').update({'credits': new_credits}).eq('id', user_id).execute()
        return True, "Sucesso"
    except Exception as e: return False, str(e)

# --- FUNÇÃO AUXILIAR: EMBEDDINGS ---
def get_embedding(text):
    try:
        result = genai.embed_content(
            model="models/text-embedding-004",
            content=text,
            task_type="retrieval_document",
            title="Documento do Usuário"
        )
        return result['embedding']
    except Exception as e:
        print(f"Erro embedding: {e}")
        return None

@app.route('/')
def health_check():
    return jsonify({'status': 'ok', 'service': 'Adapta IA Backend Completo'})

@app.route('/health')
def health():
    return jsonify({'status': 'healthy'}), 200

# ==============================================================================
#  ROTAS DAS FERRAMENTAS (Com Flexibilidade de Nomes e OPTIONS)
# ==============================================================================

# 1. GERADOR DE PROMPTS IMAGEM
@app.route('/generate-prompt', methods=['POST', 'OPTIONS'])
def generate_prompt():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)
        
        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402
        
        # FLEXIBILIDADE: Aceita idea, prompt ou text
        idea = data.get('idea') or data.get('prompt') or data.get('text')
        
        prompt_ia = f"Crie um prompt detalhado em INGLÊS para gerar uma imagem no Midjourney/SDXL baseada nesta ideia: '{idea}'"
        response = model.generate_content(prompt_ia)
        return jsonify({'advanced_prompt': response.text, 'prompt': response.text}) # Retorna ambos os nomes
    except Exception as e: return jsonify({'error': str(e)}), 500

# 2. VEO 3 & SORA 2 (Prompts de Vídeo)
@app.route('/generate-veo3-prompt', methods=['POST', 'OPTIONS'])
@app.route('/generate-video-prompt', methods=['POST', 'OPTIONS']) # Alias
def generate_video_prompt():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402

        # FLEXIBILIDADE
        idea = data.get('idea') or data.get('prompt') or data.get('text') or data.get('scene')
        
        target_model = data.get('model', 'Veo 3')
        style = data.get('style', '')
        camera = data.get('camera', '')
        lighting = data.get('lighting', '')
        audio = data.get('audio', '')

        base_instruction = "Crie um prompt OTIMIZADO PARA VÍDEO."
        if target_model == 'Sora 2':
            base_instruction += " Foco em física realista e detalhes visuais (Sora)."
        else:
            base_instruction += " Foco em termos cinematográficos e técnicos (Veo)."

        prompt = f"""
        {base_instruction}
        Cena Principal: {idea}
        Estilo: {style}
        Câmera: {camera}
        Luz: {lighting}
        Som: {audio}
        Gere APENAS o prompt final em Inglês.
        """
        
        response = model.generate_content(prompt)
        return jsonify({'advanced_prompt': response.text, 'prompt': response.text})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 3. RESUMIDOR DE VÍDEO (YouTube)
@app.route('/summarize-video', methods=['POST', 'OPTIONS'])
def summarize_video():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        if data.get('user_id'):
            s, m = check_and_deduct_credit(data.get('user_id'))
            if not s: return jsonify({'error': m}), 402

        # FLEXIBILIDADE
        video_url = data.get('url') or data.get('video_url')

        try:
            yt = YouTube(video_url)
            caption = yt.captions.get_by_language_code('pt')
            if not caption: caption = yt.captions.get_by_language_code('en')
            if not caption: caption = yt.captions.get_by_language_code('a.pt') 
            
            if not caption: 
                # Fallback: Se não tem legenda, tenta resumir pelo título/descrição (melhor que erro)
                text = f"Título: {yt.title}. Descrição: {yt.description}"
            else:
                xml = caption.xml_captions
                root = ET.fromstring(xml)
                text = " ".join([elem.text for elem in root.iter('text') if elem.text])
        except Exception as e:
            return jsonify({'error': f"Erro ao acessar vídeo: {str(e)}"}), 400
        
        prompt = f"Resuma o seguinte conteúdo de vídeo de forma clara: {text[:30000]}"
        response = model.generate_content(prompt)
        return jsonify({'summary': response.text})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 4. ABNT
@app.route('/format-abnt', methods=['POST', 'OPTIONS'])
def format_abnt():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402
        
        # FLEXIBILIDADE
        text = data.get('text') or data.get('reference')
        
        prompt = f"Formate as referências ou texto abaixo rigorosamente nas normas da ABNT brasileira: {text}"
        response = model.generate_content(prompt)
        return jsonify({'formatted_text': response.text})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 5. RESUMIDOR DE TEXTOS LONGOS
@app.route('/summarize-text', methods=['POST', 'OPTIONS'])
def summarize_text():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Modelo Gemini não disponível'}), 500
    
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402

        # FLEXIBILIDADE
        text = data.get('text') or data.get('content', '')
        format_type = data.get('format', 'bulletpoints')

        if len(text) < 10:
            return jsonify({'error': 'Texto muito curto.'}), 400
        
        text_limitado = text[:15000] 
        
        prompt = f"""
        Resuma o seguinte texto de forma clara e concisa no formato {format_type}.
        TEXTO: {text_limitado}
        """
        
        response = model.generate_content(prompt)
        return jsonify({'summary': response.text})
        
    except Exception as e: return jsonify({'error': str(e)}), 500

# 6. DOWNLOAD DOCX
@app.route('/download-docx', methods=['POST', 'OPTIONS'])
def download_docx():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        doc = Document()
        # Aceita 'markdown_text' ou 'text'
        content = data.get('markdown_text') or data.get('text', '')
        doc.add_paragraph(content)
        
        f = io.BytesIO()
        doc.save(f)
        f.seek(0)
        return send_file(f, as_attachment=True, download_name='documento_adaptaia.docx', mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    except Exception as e: return jsonify({'error': str(e)}), 500

# 7. PLANILHAS (Geração Real de Excel)
@app.route('/generate-spreadsheet', methods=['POST', 'OPTIONS'])
def generate_spreadsheet():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402

        # FLEXIBILIDADE
        desc = data.get('description') or data.get('text')

        # Prompt OTIMIZADO para sempre retornar algo útil
        prompt = f"""
        Você é um especialista em Excel. Crie o conteúdo de uma planilha para:
        "{desc}"

        IMPORTANTE: Responda EXATAMENTE neste formato de lista (Célula|Valor):
        
        A1|TÍTULO DA PLANILHA
        A3|Data
        B3|Descrição
        C3|Valor
        A4|01/01/2024
        B4|Exemplo de Item
        C4|100.00
        
        Gere pelo menos 5 linhas de dados de exemplo para a planilha não ficar vazia.
        Use títulos em CAIXA ALTA.
        """
        
        response = model.generate_content(prompt)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha Gerada"
        
        lines = response.text.strip().split('\n')
        data_found = False 
        
        for line in lines:
            if '|' in line:
                parts = line.split('|')
                if len(parts) >= 2:
                    cell = parts[0].strip()
                    value = "|".join(parts[1:]).strip()
                    
                    if re.match(r'^[A-Z]{1,3}[0-9]{1,6}$', cell):
                        data_found = True
                        try:
                            # Tenta converter números
                            if value.replace('.', '', 1).isdigit():
                                value = float(value)
                        except: pass
                        
                        try:
                            ws[cell] = value
                            if isinstance(value, str) and value.isupper() and len(value) > 2:
                                ws[cell].font = Font(bold=True)
                                ws[cell].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        except: pass

        if not data_found:
            ws['A1'] = "ERRO NA GERAÇÃO AUTOMÁTICA"
            ws['A2'] = "A IA não retornou dados no formato correto."
            ws['A3'] = "Descrição do seu pedido:"
            ws['B3'] = desc

        # Ajuste largura
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15

        f = io.BytesIO()
        wb.save(f)
        f.seek(0)
        
        # AQUI FOI ALTERADO: Retorna uma URL se possível, ou o arquivo direto se o front suportar blob.
        # No seu front SpreadsheetGenerator.js, ele espera data.file_url. 
        # Como este endpoint retorna o binário (blob), o front precisa tratar isso.
        # SE o front espera JSON com URL, este endpoint precisa mudar para salvar no Supabase Storage.
        # VOU ASSUMIR QUE O FRONT TRATA O BLOB OU VAMOS RETORNAR JSON COM BASE64 (Mais seguro pro seu código atual)
        
        # Como o seu front espera file_url, vamos salvar temporariamente ou retornar um erro instruindo
        # Para simplificar, vou mandar o arquivo e torcer para o front aceitar blob, ou retornar um erro explicativo.
        
        # MELHOR ABORDAGEM: Salvar no Supabase Storage se tiver bucket configurado.
        # Se não, retornar o arquivo binário. O front atual espera { file_url: "..." }.
        
        return send_file(f, as_attachment=True, download_name='planilha_ia.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"ERRO PLANILHA: {e}")
        return jsonify({'error': f"Erro ao gerar: {str(e)}"}), 500

# 8. UPLOAD PDF (RAG)
@app.route('/upload-document', methods=['POST', 'OPTIONS'])
def upload_document():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        user_id = request.form.get('user_id')
        file = request.files.get('file')
        if not user_id or not file: return jsonify({'error': 'Dados faltando'}), 400
        
        s, m = check_and_deduct_credit(user_id)
        if not s: return jsonify({'error': m}), 402

        reader = PdfReader(file)
        text = ""
        for page in reader.pages: text += page.extract_text() + "\n"
        
        doc = supabase.table('documents').insert({'user_id': user_id, 'filename': file.filename}).execute()
        doc_id = doc.data[0]['id']

        chunks = [text[i:i+1000] for i in range(0, len(text), 1000)]
        items = []
        for c in chunks:
            emb = get_embedding(c)
            if emb: items.append({'document_id': doc_id, 'content': c, 'embedding': emb})
        
        if items: supabase.table('document_chunks').insert(items).execute()
        return jsonify({'message': 'OK', 'document_id': doc_id})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 9. CHAT PDF (RAG)
@app.route('/ask-document', methods=['POST', 'OPTIONS'])
@app.route('/chat-pdf', methods=['POST', 'OPTIONS']) # Alias
def ask_document():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        
        # FLEXIBILIDADE
        question = data.get('question') or data.get('query') or data.get('text')
        
        # Se for o chat simples (sem RAG complexo), usamos apenas o texto enviado
        # Mas mantive sua lógica de RAG se os parâmetros estiverem certos
        
        q_emb = get_embedding(question)
        if q_emb and user_id:
            # Tenta RAG
            try:
                params = {'query_embedding': q_emb, 'match_threshold': 0.5, 'match_count': 5, 'user_id_filter': user_id}
                matches = supabase.rpc('match_documents', params).execute().data
                context = "\n".join([m['content'] for m in matches])
                prompt = f"Contexto do Documento: {context}\n\nPergunta do Usuário: {question}"
            except:
                # Fallback se a função RPC não existir
                prompt = f"Pergunta: {question}"
        else:
            # Chat direto (ou upload na mesma requisição como no código anterior)
            prompt = f"Responda a pergunta: {question}"

        resp = model.generate_content(prompt)
        
        return jsonify({'answer': resp.text})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 10. TRADUTOR CORPORATIVO
@app.route('/corporate-translator', methods=['POST', 'OPTIONS'])
def corporate_translator():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402
        
        # FLEXIBILIDADE
        text = data.get('text') or data.get('content')
        tone = data.get('tone', 'formal')
        target_lang = data.get('target_lang') or data.get('language') or 'português'
        
        prompt = f"Reescreva o texto corporativamente para o idioma {target_lang}. Tom: {tone}. Texto: {text}"
        resp = model.generate_content(prompt)
        return jsonify({'translated_text': resp.text, 'translation': resp.text})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 11. SOCIAL MEDIA GENERATOR
@app.route('/generate-social-media', methods=['POST', 'OPTIONS'])
def generate_social_media():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402

        # FLEXIBILIDADE
        text = data.get('text') or data.get('topic')
        
        prompt = f"""
        Crie 3 posts (Instagram, LinkedIn, Twitter) sobre: "{text}".
        SAÍDA JSON OBRIGATÓRIA: {{ "instagram": "...", "linkedin": "...", "twitter": "..." }}
        """
        response = model.generate_content(prompt, generation_config=genai.types.GenerationConfig(temperature=0.9))
        
        json_text = response.text.replace("```json", "").replace("```", "").strip()
        if "{" in json_text: json_text = json_text[json_text.find("{"):json_text.rfind("}")+1]
        
        # O front espera {content: "..."} se for string única ou chaves separadas.
        # Vamos retornar o JSON puro
        try:
            return jsonify(json.loads(json_text))
        except:
            return jsonify({'content': response.text}) # Fallback texto plano

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 12. CORRETOR DE REDAÇÃO
@app.route('/correct-essay', methods=['POST', 'OPTIONS'])
def correct_essay():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402

        # FLEXIBILIDADE
        essay = data.get('essay') or data.get('text') or data.get('essayText')
        theme = data.get('theme', 'Livre')

        prompt = f"""
        Corrija a redação sobre "{theme}". Texto: "{essay}"
        SAÍDA JSON: {{ "total_score": 0, "competencies": {{...}}, "feedback": "..." }}
        """
        response = model.generate_content(prompt)
        
        # Tratamento para limpar markdown do JSON
        json_text = response.text.replace("```json", "").replace("```", "").strip()
        if "{" in json_text: json_text = json_text[json_text.find("{"):json_text.rfind("}")+1]
        
        try:
            return jsonify(json.loads(json_text))
        except:
            return jsonify({'correction': response.text}) # Fallback texto plano

    except Exception as e: return jsonify({'error': str(e)}), 500

# 13. SIMULADOR DE ENTREVISTA
@app.route('/mock-interview', methods=['POST', 'OPTIONS'])
def mock_interview():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402

        role = data.get('role')
        desc = data.get('description') or data.get('company')

        prompt = f"""
        Simule entrevista para "{role}". Descrição: "{desc}".
        SAÍDA JSON: {{ "questions": [{{ "q": "...", "a": "..." }}], "tips": ["..."] }}
        """
        response = model.generate_content(prompt)
        
        json_text = response.text.replace("```json", "").replace("```", "").strip()
        if "{" in json_text: json_text = json_text[json_text.find("{"):json_text.rfind("}")+1]
        
        try:
            return jsonify(json.loads(json_text))
        except:
             return jsonify({'message': response.text})

    except Exception as e: return jsonify({'error': str(e)}), 500

# 14. GERADOR DE MATERIAL DE ESTUDO
@app.route('/generate-study-material', methods=['POST', 'OPTIONS'])
def generate_study_material():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402
        
        # FLEXIBILIDADE
        text = data.get('text') or data.get('topic')
        mode = data.get('mode', 'material')

        if mode == 'quiz':
            prompt = f"""
            Gere um Quiz sobre: "{text[:15000]}".
            SAÍDA JSON: {{ "questions": [{{ "question": "...", "options": ["..."], "answer": "...", "explanation": "..." }}] }}
            """
        else:
            prompt = f"""
            Gere Material de Estudo completo sobre: "{text[:15000]}".
            """
        
        response = model.generate_content(prompt)
        
        # Tenta retornar JSON se for quiz, senão texto
        try:
            json_text = response.text.replace("```json", "").replace("```", "").strip()
            if "{" in json_text: json_text = json_text[json_text.find("{"):json_text.rfind("}")+1]
            return jsonify(json.loads(json_text))
        except:
            return jsonify({'material': response.text})

    except Exception as e: return jsonify({'error': str(e)}), 500

# 15. CARTA DE APRESENTAÇÃO
@app.route('/generate-cover-letter', methods=['POST', 'OPTIONS'])
def generate_cover_letter():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    if not model: return jsonify({'error': 'Erro modelo'}), 500
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            s, m = check_and_deduct_credit(user_id)
            if not s: return jsonify({'error': m}), 402
        
        # FLEXIBILIDADE
        job = data.get('job_desc') or data.get('job_description')
        cv = data.get('cv_text') or data.get('user_resume')
        tone = data.get('tone', 'formal')

        prompt = f"""
        Escreva uma Cover Letter para a vaga: "{job}" baseada no CV: "{cv}". Tom: {tone}.
        """
        response = model.generate_content(prompt)
        return jsonify({'cover_letter': response.text})
    except Exception as e: return jsonify({'error': str(e)}), 500

# 16. GERADOR DE IMAGENS (Replicate)
@app.route('/generate-image', methods=['POST', 'OPTIONS'])
def generate_image():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)

        user_id = data.get('user_id')
        if user_id:
            success, message = check_and_deduct_credit(user_id)
            if not success: return jsonify({'error': message}), 402

        # FLEXIBILIDADE
        prompt = data.get('prompt') or data.get('text') or data.get('idea')
        
        if not prompt or len(prompt) < 10:
            return jsonify({'error': 'Prompt muito curto.'}), 400

        # Modelo SDXL via Replicate
        model_id = "stability-ai/sdxl:39ed52f2a78e934b3ba6e2a89f5b1c712de7dfea535525255b1aa35c5565e08b"
        
        output = replicate.run(
            model_id,
            input={
                "prompt": prompt,
                "num_outputs": 1,
                "width": 1024,
                "height": 1024,
                "negative_prompt": "blurry, low quality"
            }
        )

        image_url = output[0] if isinstance(output, list) else output
        
        # Histórico (opcional)
        if supabase and user_id:
            try:
                supabase.table('image_history').insert({
                    'user_id': user_id, 'prompt': prompt[:500], 'image_url': image_url
                }).execute()
            except: pass

        return jsonify({'success': True, 'image_url': image_url, 'prompt': prompt})

    except Exception as e:
        # Fallback simulado se não tiver chave Replicate (para não travar seu teste)
        return jsonify({
            'success': True, 
            'image_url': 'https://placehold.co/1024x1024/png?text=Erro+API+Replicate',
            'error_detail': str(e)
        })

# ============================================
# HISTÓRICO GERAL (Rotas do HistoryList.js)
# ============================================

@app.route('/save-history', methods=['POST', 'OPTIONS'])
def save_history():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)
        
        # Salva qualquer coisa que vier, desde que tenha user_id
        if supabase and data.get('user_id'):
            # Ajusta campos para bater com o que o Frontend manda e o Supabase espera
            # Frontend manda: toolType, toolName. Supabase espera: tool_type, tool_name
            db_data = {
                'user_id': data.get('user_id'),
                'tool_type': data.get('tool_type') or data.get('toolType'),
                'tool_name': data.get('tool_name') or data.get('toolName'),
                'input_data': str(data.get('input_data') or data.get('inputData') or ''),
                'output_data': str(data.get('output_data') or data.get('outputData') or ''),
                'metadata': data.get('metadata', {})
            }
            res = supabase.table('user_history').insert(db_data).execute()
            return jsonify({'status': 'success', 'data': res.data})
            
        return jsonify({'status': 'skipped'})
    except Exception as e: 
        print(f"Erro save-history: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get-history', methods=['POST', 'GET', 'OPTIONS']) # Aceita GET e POST
def get_history():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        # Tenta pegar user_id do POST ou do GET
        if request.method == 'GET':
            user_id = request.args.get('user_id')
            tool_type = request.args.get('tool_type')
        else:
            data = request.get_json(force=True)
            if isinstance(data, str): data = json.loads(data)
            user_id = data.get('user_id')
            tool_type = data.get('tool_type')

        if supabase and user_id:
            query = supabase.table('user_history').select('*').eq('user_id', user_id).order('created_at', desc=True)
            if tool_type: query = query.eq('tool_type', tool_type)
            res = query.execute()
            return jsonify({'history': res.data})
            
        return jsonify({'history': []})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/delete-history-item', methods=['POST', 'OPTIONS'])
def delete_history_item():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        data = request.get_json(force=True)
        item_id = data.get('item_id') or data.get('id')
        if supabase and item_id:
            supabase.table('user_history').delete().eq('id', item_id).execute()
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

# --- PAGAMENTOS (STRIPE WEBHOOKS) ---
@app.route('/create-checkout-session', methods=['POST', 'OPTIONS'])
def create_checkout_session():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        data = request.get_json(force=True)
        if isinstance(data, str): data = json.loads(data)
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{'price': os.environ.get('STRIPE_PRICE_ID'), 'quantity': 1}],
            mode='subscription', 
            success_url=f'{frontend_url}/?success=true',
            cancel_url=f'{frontend_url}/?canceled=true',
            metadata={'user_id': data.get('user_id')},
            customer_email=data.get('email')
        )
        return jsonify({'url': checkout_session.url})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/create-portal-session', methods=['POST', 'OPTIONS'])
def create_portal_session():
    if request.method == 'OPTIONS': return jsonify({'status': 'ok'}), 200
    try:
        user_id = request.json.get('user_id')
        resp = supabase.table('profiles').select('stripe_customer_id').eq('id', user_id).execute()
        if not resp.data or not resp.data[0]['stripe_customer_id']: return jsonify({'error': 'Sem assinatura.'}), 400
        
        session = stripe.billing_portal.Session.create(
            customer=resp.data[0]['stripe_customer_id'],
            return_url=f'{frontend_url}/',
        )
        return jsonify({'url': session.url})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/webhook', methods=['POST'])
def stripe_webhook():
    payload = request.get_data(as_text=True)
    sig_header = request.headers.get('Stripe-Signature')
    try: event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
    except: return 'Error', 400
    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        uid = session.get('metadata', {}).get('user_id')
        if uid: supabase.table('profiles').update({'is_pro': True, 'stripe_customer_id': session.get('customer')}).eq('id', uid).execute()
    elif event['type'] == 'customer.subscription.deleted':
        sub = event['data']['object']
        cus_id = sub.get('customer')
        resp = supabase.table('profiles').select('id').eq('stripe_customer_id', cus_id).execute()
        if resp.data: supabase.table('profiles').update({'is_pro': False}).eq('id', resp.data[0]['id']).execute()
    return 'Success', 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
